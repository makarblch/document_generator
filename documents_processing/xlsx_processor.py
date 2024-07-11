import re
from openpyxl import load_workbook
from config import REGEX_TAG
from tools.utilities import create_input_path, create_output_path, clear_string, make_set


def open_workbook(template_path: str):
    try:
        wb = load_workbook(create_input_path(template_path))
        return wb
    except Exception as e:
        print(f"Ошибка при загрузке файла Excel: {e}")
        return None


def replace_in_cell(cell, values):
    text = cell.value
    if not text or not isinstance(text, str):
        return

    keys = re.findall(REGEX_TAG, text)
    for key in keys:
        if key in values:
            replacement = values[key]
            placeholder = f'{{{{{key}}}}}'

            # Check if the replacement value is a number
            if isinstance(replacement, (int, float)):
                text = text.replace(placeholder, str(replacement))
            else:
                text = text.replace(placeholder, replacement)

            text = re.sub(rf'\s*\(.*?\)', '', text)
            text = re.sub(r'_+', lambda m: '_' * (len(m.group()) - len(replacement)), text, 1)

    text = re.sub(r'\([^()]*?(Ф\.И\.О|Дата|дата)[^()]*?\)', '', text)
    text = clear_string(text)

    # Set the cell value to the appropriate type
    if text.isdigit():
        cell.value = int(text)
    elif text.replace('.', '', 1).isdigit():
        cell.value = float(text)
    else:
        cell.value = text


def fill_xlsx_template(template_path, output_path, values):
    wb = open_workbook(template_path)
    if not wb:
        return

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                replace_in_cell(cell, values)

    out = create_output_path(output_path)
    wb.save(out)


def get_all_tags_xlsx(template_path):
    wb = open_workbook(template_path)
    if not wb:
        return []

    tags = set()
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    tags.update(re.findall(REGEX_TAG, cell.value))

    return tags
